
import json
import os
import socket
import subprocess
import logging
import datetime 
import openpyxl as Py_Excel





#>>>>>>>>>>>>>>>>   Connection Handlers
  

def Establishing_connection():
    
    global conn, Test_type, Bandwidth_Range, Recived_test_output, iterations,Iperf_tool_path
    global Current_Test_Bandwidths, Ipv4_adress, RSSI_Values, DUT_type, Current_time, Distance
    
    # Storing Today's date and Current time 
    Current_time=datetime.datetime.now().strftime("%d_%b_%Y__%H-%M-%S")
    
    
    # Choose a port number
    port = 666
    # Default Iteration
    iterations=3
    #Output Storing Here
    Recived_test_output={}
    
    # RSSI Values 
    RSSI_Values={}
    
    
    Current_Test_Bandwidths=[]
    
    # Get the local machine name
    host = socket.gethostname()
    Ipv4_adress= socket.gethostbyname(host)

    print(f"\n\t Host IP Address >> {Ipv4_adress}\n")
    
    # Ipfer tool path
    Iperf_tool_path = os.path.join(os.getcwd() , "Iperf_Tool\iperf.exe")
    
    # Create Script_Logs Folder if not present
    if not os.path.exists("Logs"):
            os.makedirs("Logs")
    

    # Create a socket object
    s = socket.socket()
    
    # Bind the socket to the host and port
    s.bind((Ipv4_adress, port))
    
    # Listen for incoming connections
    s.listen(1)
    
    # Accept a connection
    print('\nWaiting for a connection...')
    conn, addr = s.accept()
    print(f'\nConnected by {addr}')
    
    
    # Distance
    Distance = conn.recv(1024).decode()
    print("Distance ",Distance)
    
    # DUT Type storing
    DUT_type = conn.recv(1024).decode()
    print("DUT_type ",DUT_type)


    # Test Type storing
    Test_type = conn.recv(1024).decode()
    print("Test_type ",Test_type)

    # Bandwidth_Range storing
    Bandwidth_Range = int(conn.recv(1024).decode())
    Bandwidth_Range = "Basic Test 5 Mbits - 25 Mbis" if Bandwidth_Range == 1 else "Moderate Test 10 Mbits - 50 Mbis" if Bandwidth_Range == 2 else "Extreme Test 20 Mbits - 100 Mbis"

    print("User Selected  ",Bandwidth_Range)
    
    # Configure the logging settings
    try:
        logging.basicConfig(filename=f'Logs/Primary {DUT_type} {Test_type} {Current_time}.log', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
    except:
        print("\Logs Directory not found to insert Log file\n")
    # Storing Test Details
    logging.info(f'Connected by {addr} Device.')
    
    logging.info(f'Current Test Distance is {Distance} Mtr')
    
    logging.info(f'Device Under Test Type is {DUT_type}')
    

    logging.info(f'User Selected {Test_type}')
    
    
    logging.info(f'User Selected Test is {Bandwidth_Range}')
    
    
    # Confirmation for DUT wifi connection.
    Network_status = bool(int.from_bytes(conn.recv(1024), byteorder='big'))

    if Network_status:
        
        print("\nNetwork Configured Successfully\n")
        logging.info('Network Configured Successfully')
        
        if Network_configuration():
            return True
        else:
            return False                 
        
        

    else:
        print("\nNetwork Configuration Failed ...!!!\n\n\t Please Reboot DUT and try again \n")
        logging.critical('Network Configuration Failed ...!!!\n\n\t Please Reboot DUT and try again')
        return False

def Network_configuration():
    
        global DUT_IP
        
        # Recieving for the DUT Wifi IP
        
        DUT_IP = conn.recv(1024).decode()
        
        print(f" Recievd DUT IP ==> {DUT_IP}" )
        logging.info(f'Recievd DUT IP ==> {DUT_IP}')
        
        
        tries = 0
        while "192" not in DUT_IP :
            
            print(" Waiting for new IP")
            logging.debug('Waiting for new IP')
            
            DUT_IP = conn.recv(1024).decode()     
            print(f" New DUT IP ==> {DUT_IP}" )  
            logging.debug(f'New DUT IP ==> {DUT_IP}')
            
            if tries == 3:
                print("\nNetwork Configured But Failed to get DUT IP Address ...!!!\n\n\t Please Reboot DUT and try again \n")
                logging.critical('Network Configured But Failed to get DUT IP Address ...!!!\n\n\t Please Reboot DUT and try again')
                
                return False
            else:
                tries +=1 
      
        return True
            


#>>>>>>>>>>>>>>>>   Test Executors    

def Rx_test_Executor(wifi,Min_Bandwidth,Max_Bandwidth):

        logging.info(f'{wifi} Rx Test Started for Basic Test {Min_Bandwidth} - {Max_Bandwidth} Mbits')

        Iterator = Min_Bandwidth         
        

        while True:
                Current_Test_Bandwidths.append(Min_Bandwidth)
                # Sending Current test Bandwidth
                conn.send(str(Min_Bandwidth).encode())
                logging.info(f'Current test {Min_Bandwidth}M Bandwidth Sent to Secondary for the Validation.')
                
                print(f"\nRX Test Started for {Min_Bandwidth} Mbits\n")
                logging.info(f'{wifi} RX Test Started for {Min_Bandwidth} Mbits')
                
                logging.info(f'Calling Client_cmd_execution function to statr {Min_Bandwidth} Mbits 3 iteration Test.')
                Client_cmd_execution(str(Min_Bandwidth)+"M")
                
                # Terminating Subprocess 
                #Iperf_Exe_Terminal.terminate()
                logging.info(f'{Min_Bandwidth} Mbits Test Process completed')
                
                Recieving_output=json.loads(conn.recv(1024))
                print(" Recieving_output  ==> ", Recieving_output)
                logging.info(f'Recieved {Min_Bandwidth} Mbits Test output')
                
                
                Recived_test_output.update({f'{wifi} RX {Min_Bandwidth} Mbits ': Recieving_output })
                
                Test_continue = bool(int.from_bytes(conn.recv(1024),byteorder='big'))
                logging.info(f'Recieved Test_continue update is {Test_continue}')
                
                # Checking status to stop or continue test
                if not Test_continue:
                    
                    logging.info("Recieved command to Stop test and Closing Out put file to not write ")
                    
                    logging.info("Exiting From RX Test")
                    return False


                logging.info('Continuing the test ')
                
                if Min_Bandwidth < Max_Bandwidth:
                    logging.info('Sending False Command to Not stop Rx testing')
                    # sending False condition to continue
                    conn.send(int(False).to_bytes(1,byteorder='big'))
                    # print("sent False condition ")
                    Min_Bandwidth += Iterator

                else:
                    logging.info('Sending True Command to Stop RX Test')
                # sending True condition to Stop
                    conn.send(int(True).to_bytes(1, byteorder="big"))
                    # print("sent True condition ") 
                    return True


def Client_cmd_execution(Test_Bandwidth):
    global Iperf_Exe_Terminal

    # Below command for Cient communication
    Client_cmd = [Iperf_tool_path, '-c', DUT_IP, '-u', '-b', Test_Bandwidth, '-i', '1', '-t', '60']
    Default_iteration=1
    
    while True :    
        output_file = open("iperf_output.txt", "w")
        logging.info('Output File Created to store Iperf Terminal logs')

        Iperf_Exe_Terminal=subprocess.Popen(Client_cmd, shell=True, stdout=output_file)
        logging.info(f'Client Command Executing in the Primary System with {Client_cmd} command')        
               
    
        logging.info('Waiting for the confirmation to stop test ')
        Stop_test = bool(int.from_bytes(conn.recv(1024),byteorder='big'))

        logging.info(f"recieved command to stop test {Stop_test}")        

        Iperf_Exe_Terminal.terminate()
        logging.info('Current Test terminated ')
        
        print(f"\n RX Test {Default_iteration} Iteration completed ")
        logging.info(f'RX Test {Default_iteration} Iteration completed')
        
        logging.info('Closing Ouput File to End writing')
         # Close the output file
        output_file.close()
        
        logging.info('Reading All the Iperf terminal Output captured ')
        logging.info('Reading iperf_output text file to retriew Iperf terminal logs' )
        # Read the contents of the output file
        with open("iperf_output.txt", "r") as file:
            output = file.read()
            logging.info(output)
            
        logging.info('Reading logs completed')
        
        
        print(" Stop_test recieved ",Stop_test)
        if Stop_test:
            
            logging.info('Confirmation Recieved to Stop the test ')
            break
        
        
        logging.info('Continuing the test ')
        
        
        
        Default_iteration +=1
 

def Tx_test_Executor(wifi,Server_cmd,Min_Bandwidth,Max_Bandwidth):
        
      
        
        Iterator = Min_Bandwidth

        while Min_Bandwidth <= Max_Bandwidth: 
            
            logging.info(f'Opening iperf_output file in Writing Mode to write {Min_Bandwidth} Mbits Bandwidth  Iperf terminal Logs')
            output_file = open("iperf_output.txt", "w")
            output_error_file = open("iperf_output_error.txt", "w")
            
            # Run Server command in Primary
            Iperf_Exe_Terminal=subprocess.Popen(Server_cmd, shell=True, stdout = output_file, stderr=output_error_file)
            logging.info(f'Executing Server command in the Primary system with {Server_cmd} command')
        
            # Sendind client command to Secondary system to start TX test    

            conn.send(f'iperf -c {Ipv4_adress} -u -b {Min_Bandwidth}M -i 1 -t 60 '.encode())
            logging.info(f'Sendind client command to Secondary system to start {Min_Bandwidth}M Bandwidth TX test')
                   
            Current_Test_Bandwidths.append(Min_Bandwidth)
            # Sending Current test Bandwidth
            conn.send(str(Min_Bandwidth).encode())
            logging.info(f'Sending {Min_Bandwidth} Mbits Bandwidth for validation')
            
            print(f"\nTransmission Test Started for {Min_Bandwidth} Mbits\n")
            
            # Below code to recieve # iteration Data with Average Data
            Recieving_output=json.loads(conn.recv(1024))
            logging.info(f'Recieving {Min_Bandwidth} Mbits Test output {Recieving_output}')
            print(" Transmitted_output  ==> ", Recieving_output)
            
            Recived_test_output.update({f'{wifi} TX {Min_Bandwidth} Mbits ': Recieving_output })
            
            Min_Bandwidth += Iterator
            
            # Check to stop or to continue test
            Test_continue = bool(int.from_bytes(conn.recv(1024), byteorder='big'))
            # print("Test_continue",Test_continue)\
            logging.info('Waiting for the confirmation to stop test')
            
            if not Test_continue:
                    logging.warn("Recieved command to Stop test and Closing Out put file to not write ")
                    # Close the output file
                    output_file.close()
                    
                    # Wait for the command to complete and get its return code
                    Iperf_Exe_Terminal.terminate()
                    logging.info('Terminating Tx Test process')
                    
                    logging.info('Reading iperf_output text file to retriew Iperf terminal logs' )
                    # Read the contents of the output file
                    with open("iperf_output.txt", "r") as file:
                        output = file.read()
                        logging.info(output)

                    # Print the output
                    print("output",output)
                    logging.info('Reading logs completed')
                    logging.info("Exiting From TX Test Function")
                    return False
            
            
         
            # Wait for the command to complete and get its return code
            Iperf_Exe_Terminal.terminate()
            logging.info('Terminating Tx Test process')
        
        
           # Close the output file
            output_file.close()
            logging.info('closing output file')

            with open("iperf_output.txt", "r") as file:
                output = file.read()
                logging.info(output)

            # Print the output
            print("output",output)
            logging.info('Reading logs completed')
            
        logging.info("Exiting From TX Test Executor function")

        return True
  


#>>>>>>>>>>>>>>>>   Test Drivers   

def Rx_Testing(wifi):
    
    
    
    # Starting RX test in Secondary system by Sending Server command 
    conn.send('iperf -s -u -i 1'.encode())
    logging.info('Serever command "iperf -s -u -i 1" Sent to the Secondary')
    

    #Waitingfor the confirmation
    Recieved_confirmation=conn.recv(1024) 

    Recieved_confirmation = bool(int.from_bytes(Recieved_confirmation, byteorder='big'))
    if Recieved_confirmation:
        print("\n Server Started in DUT ")
        logging.info('Confirmation Recieved from Secondary of Server Started on DUT')
        
    else:
        
       print("\n Failed to start Server in DUT...!!! \n\t Please try again... ")
       logging.error('Failed to start Server in DUT...!!! \n Please try again...')
       
       return False
            
    
    
    # Executing Client Command from Primary
    

    if "5 Mbits - 25 Mbis" in Bandwidth_Range :


        if not Rx_test_Executor(wifi,5,25):
            
            return False
        

    elif "10 Mbits - 50 Mbis" in Bandwidth_Range :

        if not Rx_test_Executor(wifi,10,50):
            
            return False
       
        
    else:
       
       if not Rx_test_Executor(wifi,20,100):
            
            return False
        


            
    # Waiting for the confirmation
    RX_status=conn.recv(1024).decode() 
    logging.info(f'Recieved RX_status is {RX_status}')
    print("RX_status  ",RX_status)   
    
    logging.info('Returing from the Rx test Function After completing test')
    return True               

def Tx_Testing(wifi):
    
    
    logging.info(f'{wifi} TX Test Begining')
    
    #Below command for server communication
    Server_cmd = [Iperf_tool_path, '-s', '-u', '-i', '1']
  
    
    

    if "5 Mbits - 25 Mbis" in Bandwidth_Range :

        if not Tx_test_Executor(wifi,Server_cmd,5,25):
       
            return False
   
            

    elif "10 Mbits - 50 Mbis" in Bandwidth_Range :
        
        if not Tx_test_Executor(wifi,Server_cmd,10,50):
       
          return False
        
    
    else:
        
        if not Tx_test_Executor(wifi,Server_cmd,20,100):
       
           return False
        
  
        
    logging.info(" Sending command to End test")
    # Sending Command to End Test
    conn.send("End_test".encode())
    logging.info('Sending End Test command to Secondary to stop test')
    logging.info(f'{wifi} TX Test Endding')
    return True
    
def Run_Iperftool(wifi):
    global RSSI_Values
    
    
    
    if "TX_Test" in Test_type:
        
        # getiing RSSI Value  
        Before_Test_RSSI_Value = conn.recv(1024).decode()
        print(f" {wifi} RSSI Value {Before_Test_RSSI_Value}")
        RSSI_Values.update({f'{wifi} Before Test':Before_Test_RSSI_Value})
        logging.info(f'{wifi} Before Starting Test the RSSI Value is {Before_Test_RSSI_Value}')
        if Tx_Testing(wifi):
                
            logging.info('Tx Test completed Sucessfully')
            
            
        else:
            print(" Test Result is Not Good ")
            logging.info('Test Result is Not Good as Expected')
                                    
        # getiing RSSI Value  
        After_Test_RSSI_Value = conn.recv(1024).decode()
        print(f"{wifi} TX test RSSI Value ",After_Test_RSSI_Value)
        RSSI_Values.update({f'{wifi} After Test':After_Test_RSSI_Value})
        logging.info(f'{wifi} After Completing the Test the RSSI Value is {After_Test_RSSI_Value}')
                
        return 
        
    elif "RX_Test" in Test_type :
        
            
        # getiing RSSI Value  
        Before_Test_RSSI_Value = conn.recv(1024).decode()
        print(f" {wifi} RSSI Value {Before_Test_RSSI_Value}")
        RSSI_Values.update({f'{wifi} Before Test':Before_Test_RSSI_Value})
        logging.info(f'{wifi} Before Starting Test the RSSI Value is {Before_Test_RSSI_Value}')
        if Rx_Testing(wifi):
                
            logging.info('Rx Test completed Sucessfully')        
            
        else:
           
            print(" Test Result is Not Good ")
            logging.info('Test Result is Not Good as Expected')

        print(" Waiting to get RSSI value")                       
        # getiing RSSI Value  
        After_Test_RSSI_Value = conn.recv(1024).decode()
        print(f"{wifi} RX test RSSI Value ",After_Test_RSSI_Value)
        RSSI_Values.update({f'{wifi} After Test':After_Test_RSSI_Value})
        
        logging.info(f'{wifi} After Completing the Test the RSSI Value is {After_Test_RSSI_Value}')
                
        return 
                
    else:
    
   
    #------------------------------------------------------      RX  ------------------------------------------         
            
        # getiing RSSI Value  
        Before_Test_RSSI_Value = conn.recv(1024).decode()
        print(f" {wifi} RSSI Value {Before_Test_RSSI_Value}")
        RSSI_Values.update({f'{wifi} Before Test':Before_Test_RSSI_Value})
        logging.info(f'{wifi} Before Starting Test the RSSI Value is {Before_Test_RSSI_Value}')
        if Rx_Testing(wifi):
                
            logging.info(f'{wifi} Rx Test completed Sucessfully')
                       
            
        else:
            print(f"{wifi} Rx Test Result is Not Good ")
            logging.info(f"{wifi} Rx Test Result is Not Good ")

    #----------------------------------------------------------------   Tx    -----------------------------------
               


        if Tx_Testing(wifi):
                
                logging.info(f'{wifi} Tx Test completed Sucessfully')           
                
        else:
            print(f"{wifi} Tx Test Result is Not Good ")
            logging.info(f"{wifi} Tx Test Result is Not Good ")
                                    
        # getiing RSSI Value  
        After_Test_RSSI_Value = conn.recv(1024).decode()
        print(f"{wifi} TX test RSSI Value ",After_Test_RSSI_Value)
        RSSI_Values.update({f'{wifi} After Test':Before_Test_RSSI_Value})
        logging.info(f'{wifi} After Completing the Test the RSSI Value is {After_Test_RSSI_Value}')
                
        return      
 

#>>>>>>>>>>>>>>>>   Result Data Saver    

def create_excel_table(Test_data,CurrentTesttype):

    logging.info(f'Current test data contains Below Data')
    for Test_Name,Test_Values in Test_data.items():
        logging.info(f'{Test_Name} - {Test_Values}')
   
    logging.info('Inside creating Excel file')
    Test_output_Xl_file = f'Iperf {DUT_type} {Test_type} {Current_time}'
    
    # Write the headers
    headers = ['Connection Type','Mode of Connection', 'Antenna Type', 'Distance', 'RSSI',
               'Bandwidth', 'Throughput test 1 (in Mbps)',
               'Throughput test 2 (in Mbps)', 'Throughput test 3 (in Mbps)',
               'Average Throughput (in Mbps)']
    
   
    
    # Create a new workbook and select the active sheet
    workbook = Py_Excel.Workbook()  
    
        
    

    if "RX_Test" in Test_type:
        logging.info('Inside RX Test')
        if CurrentTesttype == "2.4Ghz":
            logging.info('Loading RX test data for 2.4Ghz for 1st time')
            # Create a new sheet named "RX_Test" and set it as the active sheet
            rx_sheet = workbook.active
            rx_sheet.title = "RX_Test" 
            
            rx_sheet.append([])
            rx_sheet.append(headers)     
              
        else :
            # Open the Excel file
            workbook = Py_Excel.load_workbook(f'{Test_output_Xl_file}.xlsx')
            # Access the new sheet
            rx_sheet = workbook['RX_Test']
            logging.info('Loading RX test data for 5Ghz for 2nd time')
        
     
        
    elif "TX_Test" in Test_type:
        if CurrentTesttype == "2.4Ghz":
            logging.info('Loading TX test data for 2.4Ghz for 1st time')
            # Create a new sheet named "TX_Test" and set it as the active sheet
            tx_sheet = workbook.active
            tx_sheet.title = "TX_Test"
            
            tx_sheet.append([])
            tx_sheet.append(headers)
            
        else :
            # Open the Excel file
            workbook = Py_Excel.load_workbook(f'{Test_output_Xl_file}.xlsx')
            # Access the new sheet
            tx_sheet = workbook['TX_Test']
            logging.info('Loading RX test data for 5Ghz for 2nd time')
        
        
    else:
        if CurrentTesttype == "2.4Ghz":
            # Create a new sheet named "RX_Test" and set it as the active sheet
            rx_sheet = workbook.active
            rx_sheet.title = "RX_Test"
            
            # Create a new sheet named "TX_Test"
            tx_sheet = workbook.create_sheet("TX_Test")
            
            rx_sheet.append([])
            tx_sheet.append([])
            rx_sheet.append(headers)
            tx_sheet.append(headers)
            
            logging.info('Loading RX and TX test data for 2.4Ghz for 1st time')
    
        else:
            # Open the Excel file
            workbook = Py_Excel.load_workbook(f'{Test_output_Xl_file}.xlsx')
            rx_sheet = workbook['RX_Test']
            tx_sheet = workbook['TX_Test']
            logging.info('Loading RX and TX test data for 5Ghz for 2nd time')
            
            

        
    Current_Test_Bandwidths_value_index = 0   
    Row_num=0
    Check_row=True
    for Test_Name,Test_Values in Test_data.items():
        
        if "RX" in Test_Name:
           
            if "2.4Ghz" in Test_Name:
                
                if Row_num == 0:
                    Throughput_test_1,Throughput_test_2,Throughput_test_3,Average_test = Test_Values
            
                    print(['2.4Ghz','UDP', 'PCBA', Distance+ ' mtr', RSSI_Values["2.4Ghz Before Test"],float(Current_Test_Bandwidths[Current_Test_Bandwidths_value_index]),Throughput_test_1.strip("Mbits"),Throughput_test_2.strip("Mbits"),Throughput_test_3.strip("Mbits"),Average_test.strip("Mbits") ])
                    rx_sheet.append(['2.4Ghz','UDP', 'PCBA', Distance +' mtr', "Before Test "+RSSI_Values["2.4Ghz Before Test"],Current_Test_Bandwidths[Current_Test_Bandwidths_value_index],Throughput_test_1.strip("Mbits"),Throughput_test_2.strip("Mbits"),Throughput_test_3.strip("Mbits"),Average_test.strip("Mbits") ])
                    Current_Test_Bandwidths_value_index  += 1
                elif Row_num == 1:
                    Throughput_test_1,Throughput_test_2,Throughput_test_3,Average_test = Test_Values
            
                    print(['','', '', '', RSSI_Values["2.4Ghz After Test"],Current_Test_Bandwidths[Current_Test_Bandwidths_value_index],Throughput_test_1.strip("Mbits"),Throughput_test_2.strip("Mbits"),Throughput_test_3.strip("Mbits"),Average_test.strip("Mbits") ])
                    rx_sheet.append(['','', '', '', "After Test "+RSSI_Values["2.4Ghz After Test"],Current_Test_Bandwidths[Current_Test_Bandwidths_value_index],Throughput_test_1.strip("Mbits"),Throughput_test_2.strip("Mbits"),Throughput_test_3.strip("Mbits"),Average_test.strip("Mbits") ])
                    Current_Test_Bandwidths_value_index  += 1
                    
                else:
                    Throughput_test_1,Throughput_test_2,Throughput_test_3,Average_test = Test_Values
            
                    print(['','', '', '', '',Current_Test_Bandwidths[Current_Test_Bandwidths_value_index],Throughput_test_1.strip("Mbits"),Throughput_test_2.strip("Mbits"),Throughput_test_3.strip("Mbits"),Average_test.strip("Mbits") ])
                    rx_sheet.append(['','', '', '', '',Current_Test_Bandwidths[Current_Test_Bandwidths_value_index],Throughput_test_1.strip("Mbits"),Throughput_test_2.strip("Mbits"),Throughput_test_3.strip("Mbits"),Average_test.strip("Mbits") ])
                    Current_Test_Bandwidths_value_index  += 1
                    
                        
                if Row_num!=4:
                    Row_num+=1
                
                   
                
            else:
                
                if Row_num == 0:
                    rx_sheet.append([])
                    Throughput_test_1,Throughput_test_2,Throughput_test_3,Average_test = Test_Values
            
                    print(['5Ghz','UDP', 'PCBA', Distance+' mtr', RSSI_Values["5Ghz Before Test"],Current_Test_Bandwidths[Current_Test_Bandwidths_value_index],Throughput_test_1.strip("Mbits"),Throughput_test_2.strip("Mbits"),Throughput_test_3.strip("Mbits"),Average_test.strip("Mbits") ])
                    rx_sheet.append(['5Ghz','UDP', 'PCBA', Distance +' mtr', "Before Test "+RSSI_Values["5Ghz Before Test"],Current_Test_Bandwidths[Current_Test_Bandwidths_value_index],Throughput_test_1.strip("Mbits"),Throughput_test_2.strip("Mbits"),Throughput_test_3.strip("Mbits"),Average_test.strip("Mbits") ])
                    Current_Test_Bandwidths_value_index  += 1
                elif Row_num == 1:
                    Throughput_test_1,Throughput_test_2,Throughput_test_3,Average_test = Test_Values
            
                    print(['','', '', '', RSSI_Values["5Ghz After Test"],Current_Test_Bandwidths[Current_Test_Bandwidths_value_index],Throughput_test_1.strip("Mbits"),Throughput_test_2.strip("Mbits"),Throughput_test_3.strip("Mbits"),Average_test.strip("Mbits") ])
                    rx_sheet.append(['','', '', '', "After Test "+RSSI_Values["2.4Ghz After Test"],Current_Test_Bandwidths[Current_Test_Bandwidths_value_index],Throughput_test_1.strip("Mbits"),Throughput_test_2.strip("Mbits"),Throughput_test_3.strip("Mbits"),Average_test.strip("Mbits") ])
                    Current_Test_Bandwidths_value_index  += 1
                else:
                    Throughput_test_1,Throughput_test_2,Throughput_test_3,Average_test = Test_Values
            
                    print(['','', '', '', '',Current_Test_Bandwidths[Current_Test_Bandwidths_value_index],Throughput_test_1.strip("Mbits"),Throughput_test_2.strip("Mbits"),Throughput_test_3.strip("Mbits"),Average_test.strip("Mbits") ])
                    rx_sheet.append(['','', '', '', '',Current_Test_Bandwidths[Current_Test_Bandwidths_value_index],Throughput_test_1.strip("Mbits"),Throughput_test_2.strip("Mbits"),Throughput_test_3.strip("Mbits"),Average_test.strip("Mbits") ])
                    Current_Test_Bandwidths_value_index  += 1
                if Row_num!=4:
                    Row_num+=1
               
        
        else:
            
            if "2.4Ghz" in Test_Name:

                if Check_row:
                    Row_num=0
                    Check_row=False

                
                if Row_num == 0:
                    Throughput_test_1,Throughput_test_2,Throughput_test_3,Average_test = Test_Values
            
                    print(['2.4Ghz','UDP', 'PCBA', Distance+' mtr', RSSI_Values["2.4Ghz Before Test"],float(Current_Test_Bandwidths[Current_Test_Bandwidths_value_index]),Throughput_test_1.strip("Mbits"),Throughput_test_2.strip("Mbits"),Throughput_test_3.strip("Mbits"),Average_test.strip("Mbits") ])
                    tx_sheet.append(['2.4Ghz','UDP', 'PCBA', Distance+' mtr', "Before Test "+RSSI_Values["2.4Ghz Before Test"],Current_Test_Bandwidths[Current_Test_Bandwidths_value_index],Throughput_test_1.strip("Mbits"),Throughput_test_2.strip("Mbits"),Throughput_test_3.strip("Mbits"),Average_test.strip("Mbits") ])
                    Current_Test_Bandwidths_value_index  += 1
                elif Row_num == 1:
                    Throughput_test_1,Throughput_test_2,Throughput_test_3,Average_test = Test_Values
            
                    print(['','', '', '', RSSI_Values["2.4Ghz After Test"],Current_Test_Bandwidths[Current_Test_Bandwidths_value_index],Throughput_test_1.strip("Mbits"),Throughput_test_2.strip("Mbits"),Throughput_test_3.strip("Mbits"),Average_test.strip("Mbits") ])
                    tx_sheet.append(['','', '', '', "After Test "+RSSI_Values["2.4Ghz After Test"],Current_Test_Bandwidths[Current_Test_Bandwidths_value_index],Throughput_test_1.strip("Mbits"),Throughput_test_2.strip("Mbits"),Throughput_test_3.strip("Mbits"),Average_test.strip("Mbits") ])
                    Current_Test_Bandwidths_value_index  += 1
                    
                else:
                    Throughput_test_1,Throughput_test_2,Throughput_test_3,Average_test = Test_Values
            
                    print(['','', '', '', '',Current_Test_Bandwidths[Current_Test_Bandwidths_value_index],Throughput_test_1.strip("Mbits"),Throughput_test_2.strip("Mbits"),Throughput_test_3.strip("Mbits"),Average_test.strip("Mbits") ])
                    tx_sheet.append(['','', '', '', '',Current_Test_Bandwidths[Current_Test_Bandwidths_value_index],Throughput_test_1.strip("Mbits"),Throughput_test_2.strip("Mbits"),Throughput_test_3.strip("Mbits"),Average_test.strip("Mbits") ])
                    Current_Test_Bandwidths_value_index  += 1
                        

                if Row_num!=4:
                    Row_num+=1
                else:
                    Row_num=0
                   
                
            else:
                if Check_row:
                    Row_num=0
                    Check_row=False
                    
                if Row_num == 0:
                    tx_sheet.append([])
                    Throughput_test_1,Throughput_test_2,Throughput_test_3,Average_test = Test_Values
            
                    print(['5Ghz','UDP', 'PCBA', Distance+' mtr', RSSI_Values["5Ghz Before Test"],Current_Test_Bandwidths[Current_Test_Bandwidths_value_index],Throughput_test_1.strip("Mbits"),Throughput_test_2.strip("Mbits"),Throughput_test_3.strip("Mbits"),Average_test.strip("Mbits") ])
                    tx_sheet.append(['5Ghz','UDP', 'PCBA', Distance+' mtr', "Before Test "+RSSI_Values["5Ghz Before Test"],Current_Test_Bandwidths[Current_Test_Bandwidths_value_index],Throughput_test_1.strip("Mbits"),Throughput_test_2.strip("Mbits"),Throughput_test_3.strip("Mbits"),Average_test.strip("Mbits") ])
                    Current_Test_Bandwidths_value_index  += 1
                elif Row_num == 1:
                    Throughput_test_1,Throughput_test_2,Throughput_test_3,Average_test = Test_Values
            
                    print(['','', '', '', RSSI_Values["5Ghz After Test"],Current_Test_Bandwidths[Current_Test_Bandwidths_value_index],Throughput_test_1.strip("Mbits"),Throughput_test_2.strip("Mbits"),Throughput_test_3.strip("Mbits"),Average_test.strip("Mbits") ])
                    tx_sheet.append(['','', '', '', "After Test "+RSSI_Values["2.4Ghz After Test"],Current_Test_Bandwidths[Current_Test_Bandwidths_value_index],Throughput_test_1.strip("Mbits"),Throughput_test_2.strip("Mbits"),Throughput_test_3.strip("Mbits"),Average_test.strip("Mbits") ])
                    Current_Test_Bandwidths_value_index  += 1
                else:
                    Throughput_test_1,Throughput_test_2,Throughput_test_3,Average_test = Test_Values
            
                    print(['','', '', '', '',Current_Test_Bandwidths[Current_Test_Bandwidths_value_index],Throughput_test_1.strip("Mbits"),Throughput_test_2.strip("Mbits"),Throughput_test_3.strip("Mbits"),Average_test.strip("Mbits") ])
                    tx_sheet.append(['','', '', '', '',Current_Test_Bandwidths[Current_Test_Bandwidths_value_index],Throughput_test_1.strip("Mbits"),Throughput_test_2.strip("Mbits"),Throughput_test_3.strip("Mbits"),Average_test.strip("Mbits") ])
                    Current_Test_Bandwidths_value_index  += 1
                Row_num+=1
            
   
    # Save the workbook
    workbook.save(f'{Test_output_Xl_file}.xlsx')
    logging.info(f'Test data Saved sucessfully on {Test_output_Xl_file}.xlsx')
    Current_Test_Bandwidths.clear()

def Secondary_Logs_Reciever():

    BUFFER_SIZE = 4096  # Chunk size
    file_paths=[]

    Script_log_file_path = conn.recv(1024).decode()
    print("Script_log_file_path",Script_log_file_path)
    logging.info(f" Recieved Script Log file {Script_log_file_path}")

    Device_log_file_path = conn.recv(1024).decode()

    print("Device_log_file_path",Device_log_file_path)
    logging.info(f" Recieved Device Log file {Device_log_file_path}")

 

    file_paths.append(Script_log_file_path)
    file_paths.append(Device_log_file_path)
    # print(file_paths)

    for file_path in file_paths:
        # Receive the file size from the server
        file_size_str = conn.recv(16).decode().strip()
        print("file_size_str",file_size_str)
        # Convert the file size back to an integer
        file_size = int(file_size_str)

        # Open the file in binary mode
        with open(file_path, 'wb') as file:
            total_received = 0
            while total_received < file_size:
                # Receive a chunk of data from the server
                chunk = conn.recv(BUFFER_SIZE)
                if not chunk:
                    file.close()
                    # End of file or error, break the loop
                    break
                # Write the received chunk to the file
                file.write(chunk)
                total_received += len(chunk)

    logging.info(f" All logs file saved in to Logs directory ")


# Execution Starts Here  
    
def main():
    
   
    # Building Connection
    
    if Establishing_connection():   
        
        print("Connection Established Successfully")
        logging.info('Connection Established Successfully')   



            
        Run_Iperftool("2.4Ghz")
            # Confirmation for DUT wifi connection.
        Network_status = bool(int.from_bytes(conn.recv(1024), byteorder='big'))
        
        create_excel_table(Recived_test_output,"2.4Ghz")
        Recived_test_output.clear()


        if Network_status:
                
                if Network_configuration():
                    
                    Run_Iperftool("5Ghz")
                            
                                                   
                    print("\nAll tests completed") 
                    logging.info('All the Test Completed Sucessfully')
                    
                    create_excel_table(Recived_test_output,"5Ghz")

                    Secondary_Logs_Reciever()
                                        
                    # Closing Communication Socket Here
                    conn.close()
                    logging.info('Established connection Closed Sucessfully')
                       
                else:
                    Secondary_Logs_Reciever()

                    # Closing Communication Socket Here
                    conn.close()
                    logging.info('Established connection Closed Sucessfully')

        else:
            print("\nNetwork Configuration Failed ...!!!\n\n\t Please Reboot DUT and try again \n")
            logging.critical('Network Configuration Failed ...!!!\n\n\t Please Reboot DUT and try again')
            Secondary_Logs_Reciever()

            # Closing Communication Socket Here
            conn.close() 
            
        
                    
    else:
        print("Failed to Establish Connection  ")
        logging.info('Failed to Establish Connection')        
        
        Secondary_Logs_Reciever()

        conn.close()    



if __name__ == "__main__":

    # Call the main function
    main()

    for k,v in Recived_test_output.items():

        print(f"\n\n {k} {v}")
          
    for Test__type,RSSI in RSSI_Values.items():
 
            print(f"\n\n {Test__type} {RSSI}")
        
    
    











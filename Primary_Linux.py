import openpyxl as Py_Excel
import datetime
import socket
import logging
import json
import os
import subprocess
import time
import re



class iPerf2:


    def __init__(self) -> None:
        
        
        '''
        --------------------------------------------------------------------------

        Global Static Variables 

            
        '''
        
        #Output Storing Here
        self.Recived_test_output={}

        # RSSI Values 
        self.RSSI_Values={}
        
        
        # Storing Today's date and Current time 
        self.Current_time=datetime.datetime.now().strftime("%d_%b_%Y__%H-%M-%S")
        # Choose a port number
        self.port = 666

        self.current_working_directory = os.getcwd()

        # Default Iteration
        self.iterations=3
        
        self.Throughput_test_Datagrams=[]
        
        
        self.Current_Test_Bandwidths=[]


        # Get the local machine name
        host = socket.gethostname()
        self.Ipv4_adress= socket.gethostbyname(host)

        # Ipfer tool path
        self.Iperf2_tool_path = os.path.join( self.current_working_directory , r"Iperf_Tool/iperf")

        if not os.path.exists(self.Iperf2_tool_path):
            print("\n\n\t Iperf Tool Not Exist ....!!!\n\n")


        # Create Script_Logs Folder if not present
        if not os.path.exists("logs"):
            os.makedirs("logs")
        
        # Create tmp Folder if not present
        if not os.path.exists("tmp"):
            os.makedirs("tmp")
    
    def Socket_Establisher(self):

        # Create a socket object
        self.Primary_Socket = socket.socket()


        # Bind the socket to the host and port
        self.Primary_Socket.bind((self.Ipv4_adress, self.port))

        # Listen for incoming connections
        self.Primary_Socket.listen(1)

        print(f"\n\t Host IP Address >> {self.Ipv4_adress}\n")
        
        # Accept a connection
        print('\nWaiting for a connection...')
        self.socket_conn, self.socket_addr = self.Primary_Socket.accept()
        print(f'\nConnected by {self.socket_addr}')

    def Connection_Establishing(self):
         

        self.Recieved_Usr_Input = json.loads(self.socket_conn.recv(1024))


        self.Distance = str(self.Recieved_Usr_Input[0])
        print(f"Distance {self.Distance}")
        
        # DUT Type storing
        self.DUT_type = self.Recieved_Usr_Input[1]
        print(f"DUT_type {self.DUT_type}")


        # Test Type storing
        self.Test_type = self.Recieved_Usr_Input[2]
        print(f"Test_type {self.Test_type}")
        
        # Test Dartion time storing
        self.Test_Duration = str(self.Recieved_Usr_Input[3])
        print(f"Test iteration {self.Test_Duration} sec")

        # Bandwidth_Range storing
        self.Bandwidth_Range = str(self.Recieved_Usr_Input[4])
        #self.Bandwidth_Range = "Basic Test 5 Mbits - 25 Mbis" if Bandwidth_Range == 1 else "Moderate Test 10 Mbits - 50 Mbis" if Bandwidth_Range == 2 else "Extreme Test 20 Mbits - 100 Mbis"
        print(f"User Selected  {self.Bandwidth_Range}")


         # Configure the logging settings
        try:
            logging.basicConfig(filename=f'logs/Primary {self.DUT_type} {self.Test_type} {self.Current_time}.log', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
        except:
            print("\n Logs Directory not found to insert Log file\n")
            return False
            
        
        # Storing Test Details
        logging.info(f'[Connection_Establishing] Connected by {self.socket_addr} Device.')
        
        logging.info(f'[Connection_Establishing] Current Test Distance is {self.Distance} Mtr')
        
        logging.info(f'[Connection_Establishing] Device Under Test Type is {self.DUT_type}')
        

        logging.info(f'[Connection_Establishing] User Selected {self.Test_type}')
        
        logging.info(f'[Connection_Establishing] User Entered {self.Test_Duration} seconds for test duration')
        
        
        logging.info(f'[Connection_Establishing] User Selected Test is {self.Bandwidth_Range}')
        
        return True
            
    def DUT_Network_config_Validator(self):

         # Confirmation for DUT wifi connection.
        Network_status = bool(int.from_bytes(self.socket_conn.recv(1024), byteorder='big'))

        if Network_status:
            
            print("\nNetwork Configured Successfully\n")
            logging.info('[DUT_Network_config_Validator] Network Configured Successfully')
            
            # Recieving for the DUT Wifi IP
            self.DUT_IP = self.socket_conn.recv(1024).decode()               
            
            print(f" Recievd DUT IP ==> {self.DUT_IP}" )
            logging.info(f'[DUT_Network_config_Validator] Recievd DUT IP ==> {self.DUT_IP}')
            
            return True

        else:
            print("\nNetwork Configuration Failed ...!!!\n\n\t Please Reboot DUT and try again \n")
            logging.critical('[DUT_Network_config_Validator] Network Configuration Failed ...!!!\n\n\t Please Reboot DUT and try again')
            return False      

class Data_Saver(iPerf2):
    

    def Create_Result_Excel(self,Test_data,Test_type,creating_New_file,Current_Test_Bandwidths):
        
        
        for Test_Name,Test_Values in Test_data.items():
             #print(f'{Test_Name} - {Test_Values}')
             logging.info(f"[Create_Result_Excel] {Test_Name} - {Test_Values} ")
             
        

        Test_output_Xl_file = f'Iperf {self.DUT_type} {Test_type} {self.Current_time}'
        
        # Write the headers
        headers = ['Connection Type','Mode of Connection', 'Antenna Type', 'Distance', 'RSSI',
                'Bandwidth', 
                'Throughput test 1 (in Mbps)','Throughput test 1 Lost/Total Datagrams',
                'Throughput test 2 (in Mbps)','Throughput test 2 Lost/Total Datagrams',
                'Throughput test 3 (in Mbps)','Throughput test 3 Lost/Total Datagrams',
                'Average Throughput (in Mbps)']
        
    
        
        # Create a new workbook and select the active sheet
        workbook = Py_Excel.Workbook()  
        
        
        
        if creating_New_file:
            
            
            if "RX_Test" in Test_type:
                    # print(" inside rx")
                    logging.info("[Create_Result_Excel] creating Output excel for RX Test ")
                # Create a new sheet named "RX_Test" and set it as the active sheet
                    rx_sheet = workbook.active
                    rx_sheet.title = "RX_Test" 
                    
                    rx_sheet.append([])
                    rx_sheet.append(headers) 
    
                
            elif "TX_Test" in Test_type:
            
                    # print(" inside tx")
                    logging.info("[Create_Result_Excel] creating Output excel for TX Test ")
                    # Create a new sheet named "TX_Test" and set it as the active sheet
                    tx_sheet = workbook.active
                    tx_sheet.title = "TX_Test"
                    
                    tx_sheet.append([])
                    tx_sheet.append(headers) 
                
            else:
            
                    # print(" inside else")
                    logging.info("[Create_Result_Excel] creating Output excel for FULL Test ")
                    # Create a new sheet named "RX_Test" and set it as the active sheet
                    rx_sheet = workbook.active
                    rx_sheet.title = "RX_Test"
                    
                    # Create a new sheet named "TX_Test"
                    tx_sheet = workbook.create_sheet("TX_Test")
                    
                    rx_sheet.append([])
                    tx_sheet.append([])
                    rx_sheet.append(headers)
                    tx_sheet.append(headers)

        else :
            
            if "RX_Test" in Test_type:
                # print(" inside load rx")
                logging.info("[Create_Result_Excel] loading Output excel for RX Test ")
                workbook = Py_Excel.load_workbook(f'results/{Test_output_Xl_file}.xlsx')
                # Access the new sheet
                rx_sheet = workbook['RX_Test']
                
            elif "TX_Test" in Test_type:
                # print(" inside load tx")
                logging.info("[Create_Result_Excel] loading Output excel for TX Test ")
                # Open the Excel file
                workbook = Py_Excel.load_workbook(f'results/{Test_output_Xl_file}.xlsx')
                # Access the new sheet
                tx_sheet = workbook['TX_Test']
                
            else:
                # print(" inside load else")
                logging.info("[Create_Result_Excel] loading Output excel for Full Test ")
                # Open the Excel file
                workbook = Py_Excel.load_workbook(f'results/{Test_output_Xl_file}.xlsx')
                rx_sheet = workbook['RX_Test']
                tx_sheet = workbook['TX_Test']


        if Test_Values == None:
            
            if "TX" in Test_Name:
                

                print(['','', '', '', self.RSSI_Values["After Test"]])
                tx_sheet.append(['','', '', '', "After Test "+self.RSSI_Values["After Test"]])
                logging.info('[Create_Result_Excel] For TX Only Appended After test RSSI Values ')
                                    
                self.RSSI_Values.clear()

                return

            else:
                
                print(['','', '', '', self.RSSI_Values["After Test"]])
                rx_sheet.append(['','', '', '', "After Test "+self.RSSI_Values["After Test"]])
                logging.info('[Create_Result_Excel] For RX Only Appended After test RSSI Values ')               
                self.RSSI_Values.clear()

                return
                

        if "RX" in Test_Name: 
             Data_Sheet = rx_sheet

             Throughput_test_1,Throughput_test_Datagrams_1,Throughput_test_2,Throughput_test_Datagrams_2,Throughput_test_3,Throughput_test_Datagrams_3,Average_test = Test_Values
             logging.info('[Create_Result_Excel] Extracted All the values in RX Test_Values')

        else :
                Data_Sheet = tx_sheet

                Throughput_test_1,Throughput_test_2,Throughput_test_3,Average_test = Test_Values

                
            
                # print(self.Throughput_test_Datagrams)
                Throughput_test_Datagrams_1 = self.Throughput_test_Datagrams.pop(0)
                Throughput_test_Datagrams_2 = self.Throughput_test_Datagrams.pop(0)
                Throughput_test_Datagrams_3 = self.Throughput_test_Datagrams.pop(0)

                logging.info('[Create_Result_Excel] Extracted All the values in TX Test_Values and Throughput_test_Datagrams')


         
        if "2.4Ghz" in Test_Name:
            

                if "Before Test" in self.RSSI_Values:
                    logging.info(['2.4Ghz','UDP', 'PCBA', self.Distance +' mtr', "Before Test "+self.RSSI_Values["Before Test"],Current_Test_Bandwidths,Throughput_test_1,Throughput_test_Datagrams_1,Throughput_test_2,Throughput_test_Datagrams_2,Throughput_test_3,Throughput_test_Datagrams_3,Average_test])
                    #print(['2.4Ghz','UDP', 'PCBA', self.Distance +' mtr', "Before Test "+self.RSSI_Values["Before Test"],Current_Test_Bandwidths,Throughput_test_1,Throughput_test_Datagrams_1,Throughput_test_2,Throughput_test_Datagrams_2,Throughput_test_3,Throughput_test_Datagrams_3,Average_test])
                    Data_Sheet.append(['2.4Ghz','UDP', 'PCBA', self.Distance +' mtr', "Before Test "+self.RSSI_Values["Before Test"],Current_Test_Bandwidths,Throughput_test_1,Throughput_test_Datagrams_1,Throughput_test_2,Throughput_test_Datagrams_2,Throughput_test_3,Throughput_test_Datagrams_3,Average_test])
                    
                    self.RSSI_Values.clear()
                    
                elif "After Test" in self.RSSI_Values:
                    logging.info(['','', '', '', self.RSSI_Values["After Test"],Current_Test_Bandwidths,Throughput_test_1,Throughput_test_Datagrams_1,Throughput_test_2,Throughput_test_Datagrams_2,Throughput_test_3,Throughput_test_Datagrams_3,Average_test])
                    #print(['','', '', '', self.RSSI_Values["After Test"],Current_Test_Bandwidths,Throughput_test_1,Throughput_test_Datagrams_1,Throughput_test_2,Throughput_test_Datagrams_2,Throughput_test_3,Throughput_test_Datagrams_3,Average_test])
                    Data_Sheet.append(['','', '', '', "After Test "+self.RSSI_Values["After Test"],Current_Test_Bandwidths,Throughput_test_1,Throughput_test_Datagrams_1,Throughput_test_2,Throughput_test_Datagrams_2,Throughput_test_3,Throughput_test_Datagrams_3,Average_test])
                    
                    self.RSSI_Values.clear()
                else:
                
                    logging.info(['','', '', '', Current_Test_Bandwidths,Throughput_test_1,Throughput_test_Datagrams_1,Throughput_test_2,Throughput_test_Datagrams_2,Throughput_test_3,Throughput_test_Datagrams_3,Average_test])
                    #print(['','', '', '', '',Current_Test_Bandwidths,Throughput_test_1,Throughput_test_Datagrams_1,Throughput_test_2,Throughput_test_Datagrams_2,Throughput_test_3,Throughput_test_Datagrams_3,Average_test])
                    Data_Sheet.append(['','', '', '', '',Current_Test_Bandwidths,Throughput_test_1,Throughput_test_Datagrams_1,Throughput_test_2,Throughput_test_Datagrams_2,Throughput_test_3,Throughput_test_Datagrams_3,Average_test])
        
        else:
                
                if "Before Test" in self.RSSI_Values:
                    
                    Data_Sheet.append([])
                    logging.info(['5Ghz','UDP', 'PCBA', self.Distance +' mtr', "Before Test "+self.RSSI_Values["Before Test"],Current_Test_Bandwidths,Throughput_test_1,Throughput_test_Datagrams_1,Throughput_test_2,Throughput_test_Datagrams_2,Throughput_test_3,Throughput_test_Datagrams_3,Average_test])
                    #print(['5Ghz','UDP', 'PCBA', self.Distance +' mtr', "Before Test "+self.RSSI_Values["Before Test"],Current_Test_Bandwidths,Throughput_test_1,Throughput_test_Datagrams_1,Throughput_test_2,Throughput_test_Datagrams_2,Throughput_test_3,Throughput_test_Datagrams_3,Average_test])
                    Data_Sheet.append(['5Ghz','UDP', 'PCBA', self.Distance +' mtr', "Before Test "+self.RSSI_Values["Before Test"],Current_Test_Bandwidths,Throughput_test_1,Throughput_test_Datagrams_1,Throughput_test_2,Throughput_test_Datagrams_2,Throughput_test_3,Throughput_test_Datagrams_3,Average_test])
                    
                    self.RSSI_Values.clear()
                elif "After Test" in self.RSSI_Values:
                    logging.info(['','', '', '', "After Test "+self.RSSI_Values["After Test"],Current_Test_Bandwidths,Throughput_test_1,Throughput_test_Datagrams_1,Throughput_test_2,Throughput_test_Datagrams_2,Throughput_test_3,Throughput_test_Datagrams_3,Average_test])
                    #print(['','', '', '', self.RSSI_Values["After Test"],Current_Test_Bandwidths,Throughput_test_1,Throughput_test_Datagrams_1,Throughput_test_2,Throughput_test_Datagrams_2,Throughput_test_3,Throughput_test_Datagrams_3,Average_test])
                    Data_Sheet.append(['','', '', '', "After Test "+self.RSSI_Values["After Test"],Current_Test_Bandwidths,Throughput_test_1,Throughput_test_Datagrams_1,Throughput_test_2,Throughput_test_Datagrams_2,Throughput_test_3,Throughput_test_Datagrams_3,Average_test])
                    
                    self.RSSI_Values.clear()
                else:
                    logging.info(['','', '', '', '',Current_Test_Bandwidths,Throughput_test_1,Throughput_test_Datagrams_1,Throughput_test_2,Throughput_test_Datagrams_2,Throughput_test_3,Throughput_test_Datagrams_3,Average_test])
                    #print(['','', '', '', '',Current_Test_Bandwidths,Throughput_test_1,Throughput_test_Datagrams_1,Throughput_test_2,Throughput_test_Datagrams_2,Throughput_test_3,Throughput_test_Datagrams_3,Average_test])
                    Data_Sheet.append(['','', '', '', '',Current_Test_Bandwidths,Throughput_test_1,Throughput_test_Datagrams_1,Throughput_test_2,Throughput_test_Datagrams_2,Throughput_test_3,Throughput_test_Datagrams_3,Average_test])
        
        
        # Save the workbook
        workbook.save(f'results/{Test_output_Xl_file}.xlsx')
        
        self.Recived_test_output.clear()
        logging.info('[Create_Result_Excel] Data Saved in the Excel file and cleared Test output')

    def Secondary_Logs_Reciever(self):

        self.LOG_BUFFER_SIZE = 4096  # Chunk size
        self.Secondary_logfile_paths=[]

        Script_log_file_path = self.socket_conn.recv(1024).decode()
        print("Script_log_file_path",Script_log_file_path)
        logging.info(f"[Secondary_Logs_Reciever] Recieved Script Log file {Script_log_file_path}")

        Device_log_file_path = self.socket_conn.recv(1024).decode()

        print("Device_log_file_path",Device_log_file_path)
        logging.info(f"[Secondary_Logs_Reciever] Recieved Device Log file {Device_log_file_path}")

    

        self.Secondary_logfile_paths.append(Script_log_file_path)
        self.Secondary_logfile_paths.append(Device_log_file_path)
        # print(file_paths)

        for file_path in self.Secondary_logfile_paths:
            # Receive the file size from the server
            file_size_str = self.socket_conn.recv(16).decode().strip()
            print("Recieved file_size_str",file_size_str)
            # Convert the file size back to an integer
            file_size = int(file_size_str)

            # Open the file in binary mode
            with open(file_path, 'wb') as file:
                total_received = 0
                while total_received < file_size:
                    # Receive a chunk of data from the server
                    chunk = self.socket_conn.recv(self.LOG_BUFFER_SIZE)
                    if not chunk:
                        file.close()
                        # End of file or error, break the loop
                        break
                    # Write the received chunk to the file
                    file.write(chunk)
                    total_received += len(chunk)

        logging.info("[Secondary_Logs_Reciever] All logs file saved in to Logs directory ")

class Test_Drivers(Data_Saver):

    def __init__(self,) -> None:

        super().__init__()
                
        self.check_RSSI = True

        self.TX_Test_Datagrams=[]
        self.DefaultTXDatagrams = ["None","None","None"]
        self.Average_data_finder_Pattern=r' 0\.0-\d+\.'
        
        self.DataGrams_Pattern_1 =  r'\d+/\s+\d+ \(\d+%\)'
        self.DataGrams_Pattern_2 =  r'\d+/\s+\d+ \(\d+\.\d+%\)'
        self.DataGrams_Pattern_3 =  r'\d+/+\d+ \(\d+%\)'
        self.DataGrams_Pattern_4 =  r'\d+/+\d+ \(\d+\.\d+%\)'
        self.DataGrams_Pattern_5 =  r'\d+/+\d+'

        
        
    def Rx_Driver(self,wifi):
    
        
    
        # Starting RX test in Secondary system by Sending Server command 
        self.socket_conn.send('iperf -s -u -i 1'.encode())
        logging.info('[Rx_Driver] Serever command "iperf -s -u -i 1" Sent to the Secondary')
        
        
        
        # getting RSSI Value  
        Before_Test_RSSI_Value = self.socket_conn.recv(1024).decode()
        print(f" {wifi} RSSI Value {Before_Test_RSSI_Value}")
        self.RSSI_Values.update({f'Before Test':Before_Test_RSSI_Value})
        logging.info(f'[Rx_Driver] {wifi} Before Starting Test the RSSI Value is {Before_Test_RSSI_Value}')
        

        # Waitingfor the confirmation
        Server_started_in_DUT = self.socket_conn.recv(1024) 
        Server_started_in_DUT = bool(int.from_bytes(Server_started_in_DUT, byteorder='big'))
        print(f'Server_started_in_DUT {Server_started_in_DUT}')

        if not Server_started_in_DUT:


            print("\n Failed to start Server in DUT...!!! \n\t Please try again... ")
            logging.error('[Rx_Driver] Failed to start Server in DUT...!!! \n Please try again...')
            
            return False

        else:
                        
            print("\n Server Started in DUT ")
            logging.info('[Rx_Driver] Confirmation Recieved from Secondary of Server Started on DUT')
            
                
        
        
        # Executing Client Command from Primary
        

        if "5 Mbits - 25 Mbis" in self.Bandwidth_Range :


            if not self.RX_Test(wifi,5,25):
                logging.info(f'[RX_Driver] Exiting and returing False from 5 Mbits - 25 Mbis')
                return False
            
            

        elif "10 Mbits - 50 Mbis" in self.Bandwidth_Range :

            if not self.RX_Test(wifi,10,50):
                logging.info(f'[RX_Driver] Exiting and returing False from 10 Mbits - 50 Mbis')
                return False
        
            
        else:
        
            if not self.RX_Test(wifi,20,100):
                    logging.info(f'[RX_Driver] Exiting and returing False from 20 Mbits - 100 Mbis')
                    return False
            


        logging.info('[RX_Driver] Returing from the Rx test Function After completing test')
        return True     
    
    def TX_Driver(self,wifi):

                
        logging.info(f'{wifi} TX Test Begining')
        
        #Below command for server communication
        Server_cmd = [self.Iperf2_tool_path, '-s', '-u', '-i', '1']
    
        
        

        if "5 Mbits - 25 Mbis" in self.Bandwidth_Range :

            if not self.TX_Test(wifi,Server_cmd,5,25):
                logging.info(f'[TX_Driver] Exiting and returing False from 5 Mbits - 25 Mbis')
                return False
    
                

        elif "10 Mbits - 50 Mbis" in self.Bandwidth_Range :
            
            if not self.TX_Test(wifi,Server_cmd,10,50):
                logging.info(f'[TX_Driver] Exiting and returing False from 10 Mbits - 50 Mbis')
                return False
            
        
        else:
            
            if not self.TX_Test(wifi,Server_cmd,20,100):
              logging.info(f'[TX_Driver] Exiting and returing False from 20 Mbits - 100 Mbis')
              return False
            
        logging.info(f'[TX_Driver] Exiting and returing True')
        return True

    def TX_Test(self,wifi,Server_cmd,Min_Bandwidth,Max_Bandwidth):
        

        Iterator = Min_Bandwidth
    

        while True: 
            
            
            
            logging.info(f'[TX_Test] Opening iperf_output file in Writing Mode to write {Min_Bandwidth} Mbits Bandwidth  Iperf terminal Logs')
            output_file = open("tmp/iperf_output.txt", "w")
            
            #output_error_file = open("iperf_output_error.txt", "w")
            
            # Run Server command in Primary
            Iperf_Exe_Terminal=subprocess.Popen(Server_cmd,  stdout = output_file, shell=True)
            logging.info(f'[TX_Test] Executing Server command in the Primary system with {Server_cmd} command')
        
            # Sendind client command to Secondary system to start TX test    

            # self.socket_conn.send(f'iperf -c {self.Ipv4_adress} -u -b {Min_Bandwidth}M -i 1 -t {self.Test_Duration} '.encode())
            # logging.info(f'[TX_Test] Sent iperf -c {self.Ipv4_adress} -u -b {Min_Bandwidth}M -i 1 -t {self.Test_Duration}  command to Secondary system to start {Min_Bandwidth}M Bandwidth TX test')
            
            # time.sleep(6)
            # # Sending Current test Bandwidth           
            # self.socket_conn.send(str(Min_Bandwidth).encode())
            # logging.info(f'[TX_Test] Sent {Min_Bandwidth} Mbits Bandwidth for validation')


            send_data=[f'iperf -c {self.Ipv4_adress} -u -b {Min_Bandwidth}M -i 1 -t {self.Test_Duration} ',str(Min_Bandwidth)]
            self.socket_conn.send(json.dumps(send_data).encode())

            #if self.check_RSSI and self.Test_type == "TX_Test":
            if self.check_RSSI:
                #getiing RSSI Value  
                Before_Test_RSSI_Value = self.socket_conn.recv(1024).decode()
                print(f" {wifi} RSSI Value {Before_Test_RSSI_Value}")
                self.RSSI_Values.update({f'Before Test':Before_Test_RSSI_Value})
                logging.info(f'[TX_Test] {wifi} Before Starting Test the RSSI Value is {Before_Test_RSSI_Value}')
                self.check_RSSI = False
                   
 
            print(f"\n{wifi} Transmission Test Started for {Min_Bandwidth} Mbits\n")
            logging.info(f'[TX_Test] {wifi} Transmission Test Started for {Min_Bandwidth} Mbits')
            
            
            # Below code to recieve # iteration Data with Average Data
            Recieving_output=json.loads(self.socket_conn.recv(1024))
            logging.info(f'[TX_Test] Recieved {Min_Bandwidth} Mbits Test output {Recieving_output}')
            #print(" Transmission_output  ==> ", Recieving_output)
            
            self.Recived_test_output.update({f'{wifi} TX {Min_Bandwidth} Mbits ': Recieving_output })
            

            
            # Wait for the command to complete and get its return code
            Iperf_Exe_Terminal.terminate()
            logging.info('[TX_Test] Terminating Tx Test process')
            
            time.sleep(1)
            # Close the output file
            output_file.close()
            logging.info('[TX_Test] closing output file')
            
            logging.info('[TX_Test] Reading iperf_output text file to retriew Iperf terminal logs' )
            
            
            # Read the contents of the output file        
            with open("tmp/iperf_output.txt", "r") as file:
                output = file.readlines()

                for line in output:

                    logging.info(f'[TX_Test] {line}')
                    
                    # if len(self.TX_Test_Datagrams)<=3:
                    
                    Average_Data_Found = re.search(self.Average_data_finder_Pattern,line)
                    
                    if Average_Data_Found:
                        
                        logging.info("[TX_Test] Inside Average_data_finder_Pattern ")  
                        Datagram_found = re.search(self.DataGrams_Pattern_1, line) 
                        
                        if Datagram_found :
                            
                            
                            logging.info(f"[TX_Test] Datagram found with 1st Pattern {Datagram_found[0]}")
                            self.TX_Test_Datagrams.append(Datagram_found[0])
                            #print(Datagram_found[0])
                            
                        elif re.search(self.DataGrams_Pattern_2, line):
                            
                            
                            
                            Datagram_found=re.search(self.DataGrams_Pattern_2, line)
                            logging.info(f"[TX_Test] Datagram found with 2nd Pattern {Datagram_found[0]}")
                            self.TX_Test_Datagrams.append(Datagram_found[0])
                            #print(Datagram_found[0])
                            
                        elif re.search(self.DataGrams_Pattern_3, line):
                            
                        
                            
                            Datagram_found=re.search(self.DataGrams_Pattern_3, line)
                            logging.info(f"[TX_Test] Datagram found with 3rd Pattern {Datagram_found[0]}")
                            self.TX_Test_Datagrams.append(Datagram_found[0])
                            
                            #print(Datagram_found[0])
                            
                        elif re.search(self.DataGrams_Pattern_4, line):
                            
                    
                            
                            Datagram_found=re.search(self.DataGrams_Pattern_4, line)
                            logging.info(f"[TX_Test] Datagram found with 4th Pattern {Datagram_found[0]}")
                            self.TX_Test_Datagrams.append(Datagram_found[0])
                        
                            #print(Datagram_found[0])


                        elif re.search(self.DataGrams_Pattern_5, line):
                            
                    
                            
                            Datagram_found=re.search(self.DataGrams_Pattern_5, line)
                            logging.info(f"[TX_Test] Datagram found with 5th Pattern {Datagram_found[0]}")
                            self.TX_Test_Datagrams.append(Datagram_found[0])
                        
                            #print(Datagram_found[0])
                            
                        else:
                            
                            logging.info("[TX_Test] Datagram Not found with any Pattern !!!!")
                            pass
                            


            logging.info('[TX_Test] Reading logs completed')  
            
            # if not Datagram_found : 
                
        

            logging.info(f'[TX_Test] items available in the TX datagrams list {self.TX_Test_Datagrams}')  
            
            
            if self.TX_Test_Datagrams==[] :
                
                logging.info('[TX_Test] Inside Empty TX_Test_Datagrams')
        
                self.Throughput_test_Datagrams.extend(self.DefaultTXDatagrams)
               
            elif len(self.TX_Test_Datagrams) == 2:
                
                logging.info('[TX_Test] Inside 2 items TX_Test_Datagrams')
                self.TX_Test_Datagrams.insert(0,self.DefaultTXDatagrams[0])
                logging.info(f'[TX_Test] TX_Test_Datagrams {self.TX_Test_Datagrams}')
                
                self.Throughput_test_Datagrams.extend(self.TX_Test_Datagrams)
                logging.info(f'[TX_Test] Throughput_test_Datagrams {self.Throughput_test_Datagrams}')
           
            elif len(self.TX_Test_Datagrams) == 1:
                
                logging.info('Inside 1 item TX_Test_Datagrams')
                
                self.TX_Test_Datagrams.insert(0,self.DefaultTXDatagrams[0])
                self.TX_Test_Datagrams.insert(1,self.DefaultTXDatagrams[0])
                
                logging.info(f'[TX_Test] TX_Test_Datagrams {self.TX_Test_Datagrams}')
                
                self.Throughput_test_Datagrams.extend(self.TX_Test_Datagrams)
                logging.info(f'[TX_Test] Throughput_test_Datagrams {self.Throughput_test_Datagrams}')
      
            else:
                self.Throughput_test_Datagrams.extend(self.TX_Test_Datagrams)
                logging.info(f'[TX_Test] Throughput_test_Datagrams {self.Throughput_test_Datagrams}')

            self.TX_Test_Datagrams.clear()
            
            
 
            # Check to stop or to continue test
            Test_continue = bool(int.from_bytes(self.socket_conn.recv(1024), byteorder='big'))
            # print("Test_continue",Test_continue)
            logging.info('[TX_Test] Waiting for the confirmation to stop test')
            
           
            if Min_Bandwidth  == Iterator and self.Test_type == "TX_Test" and wifi == '2.4Ghz' :
                    logging.info('[TX_Test] Calling Create Excel function to create new Excel file')
                    self.Create_Result_Excel(self.Recived_test_output,self.Test_type,True,Min_Bandwidth)
                    
                    Min_Bandwidth += Iterator
            elif Test_continue and Min_Bandwidth < Max_Bandwidth:
                    logging.info('[TX_Test] Calling Create Excel function to add in the exist Excel file')
                    self.Create_Result_Excel(self.Recived_test_output,self.Test_type,False,Min_Bandwidth)
                   
                    Min_Bandwidth += Iterator
            else:


                logging.info("[TX_Test]  Sending command to End test")
                # Sending Command to End Test
                self.socket_conn.send("End_test".encode())
                logging.info('[TX_Test] Sending End Test command to Secondary to stop test')
                logging.info(f'[TX_Test] {wifi} TX Test Endding')
                
                
                
                # getiing RSSI Value  
                After_Test_RSSI_Value = self.socket_conn.recv(1024).decode()
                print(f"{wifi} TX test RSSI Value ",After_Test_RSSI_Value)
                self.RSSI_Values.update({f'After Test':After_Test_RSSI_Value})
                logging.info(f'[TX_Test] {wifi} After Completing the Test the RSSI Value is {After_Test_RSSI_Value}')

                logging.info(f'[TX_Test] changed check_RSSI to True')
                self.check_RSSI = True
                
                self.Create_Result_Excel(self.Recived_test_output,self.Test_type,False,Min_Bandwidth)
                                        
                logging.info("[TX_Test] Exiting From TX Test Function")   
                
                logging.info("[TX_Test] Exiting From TX Test Executor function")

                return True
                

          
            if not Test_continue:
                
                    print(" Test Result is Not Good ")
                    logging.info('[TX_Test] Test Result is Not Good as Expected')
                    
                    
                    logging.info("[TX_Test] Recieved command to Stop test and Closing Out put file to not write ")
                
                    # getiing RSSI Value  
                    After_Test_RSSI_Value = self.socket_conn.recv(1024).decode()
                    print(f"{wifi} TX test RSSI Value ",After_Test_RSSI_Value)
                    self.RSSI_Values.update({f'After Test':After_Test_RSSI_Value})
                    logging.info(f'[TX_Test] {wifi} After Completing the Test the RSSI Value is {After_Test_RSSI_Value}')
                    
                    if Min_Bandwidth  == Iterator:

                        self.Recived_test_output.update({f'{wifi} TX {Min_Bandwidth} Mbits ': None }) 
                        self.Create_Result_Excel(self.Recived_test_output,self.Test_type,False,Min_Bandwidth)
                    
                    else:
                         
                        self.Create_Result_Excel(self.Recived_test_output,self.Test_type,False,Min_Bandwidth)
                                            
                    logging.info("[TX_Test] Exiting From TX Test Function")
                    return False

    def RX_Test(self,wifi,Min_Bandwidth,Max_Bandwidth):
        

        logging.info(f'[RX_Test] {wifi} Rx Test Started for Basic Test {Min_Bandwidth} - {Max_Bandwidth} Mbits')  

        Iterator = Min_Bandwidth 


        while True:
                
                # Sending Current test Bandwidth
                self.socket_conn.send(str(Min_Bandwidth).encode())
                logging.info(f'[RX_Test] Current test {Min_Bandwidth}M Bandwidth Sent to Secondary for the Validation.')
                
                print(f"\n{wifi} RX Test Started for {Min_Bandwidth} Mbits\n")
                logging.info(f'[RX_Test] {wifi} RX Test Started for {Min_Bandwidth} Mbits')
                
                logging.info(f'[RX_Test] Calling Client_cmd_execution function to statr {Min_Bandwidth} Mbits 3 iteration Test.')
                self.Client_cmd_execution(str(Min_Bandwidth)+"M")
                
                # Terminating Subprocess 
                #Iperf_Exe_Terminal.terminate()
                logging.info(f'[RX_Test] {Min_Bandwidth} Mbits Test Process completed')
                
                Recieving_output=json.loads(self.socket_conn.recv(1024))
                print(" Recieved output  ==> ", Recieving_output)
                logging.info(f'[RX_Test] Recieved {Min_Bandwidth} Mbits Test output')
                
                
                self.Recived_test_output.update({f'{wifi} RX {Min_Bandwidth} Mbits ': Recieving_output })
                
                Test_continue = bool(int.from_bytes(self.socket_conn.recv(1024),byteorder='big'))
                logging.info(f'[RX_Test] Recieved Test_continue update is {Test_continue}')


         
                # Checking status to stop or continue test
                if not Test_continue:

                    if Min_Bandwidth  == Iterator:
                         self.Create_Result_Excel(self.Recived_test_output,self.Test_type,True,Min_Bandwidth)

                    
                    logging.info("[RX_Test] Recieved command to Stop test and Closing Out put file to not write ")
                    
                    print(" Test Result is Not Good ")
                    logging.info('[RX_Test] Test Result is Not Good as Expected')
                    
                    
                    print(" Waiting to get RSSI value")                       
                    # getiing RSSI Value  
                    After_Test_RSSI_Value = self.socket_conn.recv(1024).decode()
                    print(f"{wifi} RX test RSSI Value ",After_Test_RSSI_Value)
                    self.RSSI_Values.update({f'After Test':After_Test_RSSI_Value})
                    
                    logging.info(f'[RX_Test] {wifi} After Completing the Test the RSSI Value is {After_Test_RSSI_Value}')
                    
                    if Min_Bandwidth  == Iterator:

                        self.Recived_test_output.update({f'{wifi} RX {Min_Bandwidth} Mbits ': None }) 
                        self.Create_Result_Excel(self.Recived_test_output,self.Test_type,False,Min_Bandwidth)
                    
                    else:
                         
                        self.Create_Result_Excel(self.Recived_test_output,self.Test_type,False,Min_Bandwidth)
 

                    if self.Test_type == "Full_Test":
                            TX_Before_Test_RSSI_Value = After_Test_RSSI_Value 

                            print(f" {wifi} RSSI Value {TX_Before_Test_RSSI_Value}")
                            self.RSSI_Values.update({f'Before Test':TX_Before_Test_RSSI_Value})
                            logging.info(f'[RX_Test] {wifi} Before Starting Test the RSSI Value is {TX_Before_Test_RSSI_Value}')
 
                    logging.info("[RX_Test] Exiting From RX Test")
  
                
                    return False


                
                    
                logging.info('[RX_Test] Continuing the test ')
                
                if Min_Bandwidth < Max_Bandwidth:

                    if Min_Bandwidth  == Iterator and wifi == "2.4Ghz":
                        self.Create_Result_Excel(self.Recived_test_output,self.Test_type,creating_New_file=True,Current_Test_Bandwidths=Min_Bandwidth)

                    else:
                        self.Create_Result_Excel(self.Recived_test_output,self.Test_type,False,Min_Bandwidth)


                    logging.info('[RX_Test] Sending False Command to Not stop Rx testing')
                    # sending False condition to continue
                    self.socket_conn.send(int(False).to_bytes(1,byteorder='big'))
                    # print("sent False condition ")
                    Min_Bandwidth += Iterator

                else:
                    logging.info('[RX_Test] Sending True Command to Stop RX Test')
                # sending True condition to Stop
                    self.socket_conn.send(int(True).to_bytes(1, byteorder="big"))
                    # print("sent True condition ") 
                    
                    
                    print(" Waiting to get RSSI value")                       
                    # getiing RSSI Value  
                    After_Test_RSSI_Value = self.socket_conn.recv(1024).decode()
                    print(f"{wifi} RX test RSSI Value ",After_Test_RSSI_Value)
                    self.RSSI_Values.update({f'After Test':After_Test_RSSI_Value})
                    
                    logging.info(f'{wifi} After Completing the Test the RSSI Value is {After_Test_RSSI_Value}')

                                      
                    
                    self.Create_Result_Excel(self.Recived_test_output,self.Test_type,False,Min_Bandwidth)

                                  

                    # if self.Test_type == "Full_Test":
                    #         TX_Before_Test_RSSI_Value = After_Test_RSSI_Value 

                    #         print(f" {wifi} RSSI Value {TX_Before_Test_RSSI_Value}")
                    #         self.RSSI_Values.update({f'Before Test':TX_Before_Test_RSSI_Value})
                    #         logging.info(f'[RX_Test] {wifi} Before Starting Test the RSSI Value is {TX_Before_Test_RSSI_Value}')        

                          
                    return True

    def Client_cmd_execution(self,Test_Bandwidth):
        global Iperf_Exe_Terminal

        Default_iteration = 1

        # Below command for Cient communication
        Client_cmd = [self.Iperf2_tool_path, '-c', self.DUT_IP, '-u', '-b', Test_Bandwidth, '-i', '1', '-t',self.Test_Duration]
        
        
        while True :    
            output_file = open("tmp/iperf_output.txt", "w")
            logging.info('[Client_cmd_execution] Output File Created to store Iperf Terminal logs')

            Iperf_Exe_Terminal=subprocess.Popen(Client_cmd, shell=False, stdout=output_file)
            logging.info(f'[Client_cmd_execution] Client Command Executing in the Primary System with {Client_cmd} command')        
                
        
            logging.info('[Client_cmd_execution] Waiting for the confirmation to stop test ')
            Stop_test = bool(int.from_bytes(self.socket_conn.recv(1024),byteorder='big'))

            logging.info(f"[Client_cmd_execution] recieved command to stop test {Stop_test}")        

            Iperf_Exe_Terminal.terminate()
            logging.info('[Client_cmd_execution] Current Test terminated ')
            
            print(f"\n RX Test {Default_iteration} Iteration completed ")
            logging.info(f'[Client_cmd_execution] RX Test {Default_iteration} Iteration completed')
            
            logging.info('[Client_cmd_execution] Closing Ouput File to End writing')
            # Close the output file
            output_file.close()
            
            logging.info('[Client_cmd_execution] Reading All the Iperf terminal Output captured ')
            logging.info('[Client_cmd_execution] Reading iperf_output text file to retriew Iperf terminal logs' )
            # Read the contents of the output file
            with open("tmp/iperf_output.txt", "r") as file:
                output = file.read()
                logging.info(f'[Client_cmd_execution] {output}')
                
            logging.info('[Client_cmd_execution] Reading logs completed')
            
            
            # print(" Stop_test recieved ",Stop_test)
            time.sleep(3)
            if Stop_test:
                
                logging.info('[Client_cmd_execution] Confirmation Recieved to Stop the test ')
                break
            
            
            logging.info('[Client_cmd_execution] Continuing the test ')
            
            
            
            Default_iteration +=1
    



class Execution(Test_Drivers):
            
    def Run_Iperftool(self,wifi):
  
        
        if "TX_Test" in self.Test_type:
            
            if self.TX_Driver(wifi):
                    
                logging.info('[Run_Iperftool] Tx Test completed Sucessfully')
                return True
            
            else:
                logging.info('[Run_Iperftool] returning False from TX_Test')
                return False
                
             
            
        elif "RX_Test" in self.Test_type :
            
        
            if self.Rx_Driver(wifi):
                    
                logging.info('[Run_Iperftool] Rx Test completed Sucessfully')  
                return True      
            
            else:
                logging.info('[Run_Iperftool] returning False from RX_Test')
                return False 
                    
        else:
        
    
        #------------------------------------------------------      RX  ------------------------------------------         
                
            if self.Rx_Driver(wifi): 
                    
                logging.info(f'[Run_Iperftool] {wifi} Rx Test completed Sucessfully')
                
            
        #----------------------------------------------------------------   Tx    -----------------------------------
                

            if self.TX_Driver(wifi):
                    
                    logging.info(f'[Run_Iperftool] {wifi} Tx Test completed Sucessfully')       
                    return  True
                        
                    
            else:
                print(f"{wifi} Tx Test Result is Not Good ")
                logging.info(f"[Run_Iperftool] {wifi} Tx Test Result is Not Good ")
                return  False
                                        
                 
    


# Execution Starts Here  
    
def main():
    
    iPerf = Execution()
    
    
    # Building Connection
    iPerf.Socket_Establisher()
    
    if iPerf.Connection_Establishing():   
        
        print("Connection Established Successfully")
        logging.info('[main] Connection Established Successfully')   

 
        if iPerf.DUT_Network_config_Validator():
            
            iPerf.Run_Iperftool("2.4Ghz")
    
                    
            if iPerf.DUT_Network_config_Validator():
                
                iPerf.Run_Iperftool("5Ghz")
                        
                                            
                print("\nAll tests completed") 
                logging.info('[main] All the Test Completed Sucessfully')
                
                #create_excel_table(Recived_test_output,"5Ghz")

                iPerf.Secondary_Logs_Reciever()
                                    
                # Closing Communication Socket Here
                iPerf.socket_conn.close()
                logging.info('[main] Established connection Closed Sucessfully')
                
            else:
                iPerf.Secondary_Logs_Reciever()

                # Closing Communication Socket Here
                iPerf.socket_conn.close()
                logging.info('[main] Established connection Closed Sucessfully')
                
                return False

           
        
        
        else:
            iPerf.Secondary_Logs_Reciever()

            # Closing Communication Socket Here
            iPerf.socket_conn.close()
            logging.info('[main] Established connection Closed Sucessfully')
            
            return False        
                

                    
                                
    else:
        print("Failed to Establish Connection  ")
        logging.info('[main] Failed to Establish Connection')        
        
        iPerf.Secondary_Logs_Reciever()

        # Closing Communication Socket Here
        iPerf.socket_conn.close()    







if __name__ == "__main__":

    # Call the main function
    main()
   




























